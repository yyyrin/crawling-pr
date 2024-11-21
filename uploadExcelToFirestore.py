import os
from openpyxl import load_workbook
import firebase_admin
from firebase_admin import credentials, firestore
from dotenv import load_dotenv

load_dotenv()

cred_path = os.getenv("FIREBASE_CREDENTIALS_PATH")
cred = credentials.Certificate(cred_path)
firebase_admin.initialize_app(cred)
db = firestore.client()

# 엑셀 데이터 Firestore에 저장
def upload_excel_to_Firestore(file_path):
    workbook = load_workbook(file_path)

    # Concert Info 시트 처리
    concert_sheet = workbook["Concert Info Final"]
    for row in concert_sheet.iter_rows(min_row=2, values_only=True):
        concert_id, category, concert_name, start_date, end_date, concert_venue, dup_number = row
        # concert_id = f"c{start_date.strftime('%y%m%d')}-{end_date.strftime('%y%m%d')}-{dup_number:02}"
    
        concert_data = {
            "concert_id": concert_id,
            "category": category,
            "concert_name": concert_name,
            "start_date": start_date,
            "end_date": end_date,
            "concert_venue": concert_venue,
            "dup_number": dup_number,
        }

        # Concerts 컬렉션에 저장
        db.collection("concerts").document(concert_id).set(concert_data)
        print(f"Uploaded concert: {concert_name}")

    # Setlist Info 시트 처리
    setlist_sheet = workbook["Setlist Info Final"]
    for row in setlist_sheet.iter_rows(min_row=2, values_only=True):
        concert_id, concert_name, order, track_id, title, artist, youtube_url = row

        # Songs 컬렉션에서 추가 데이터 가져오기
        song_doc = db.collection("songs").document(track_id).get()
        if song_doc.exists:
            song_data = song_doc.to_dict()
            setlist_data = {
                "order": order,
                "track_id": track_id,
                "youtube_url": youtube_url,
                "title": song_data.get("title"),
                "artist": song_data.get("artist"),
                "artist_id": song_data.get("artist_id"),
                "album_title": song_data.get("album_title"),
                "album_img": song_data.get("album_img"),
                "play_time": song_data.get("play_time"),
            }
        else:
            print(f"Warning: Song with track_id {track_id} not found.")
            setlist_data = {
                "order": order,
                "track_id": track_id,
                "youtube_url": youtube_url,
            }
        
        # Setlists 서브컬렉션에 저장
        track_doc_id = f"{concert_id}-track{order:02}"
        db.collection("concerts").document(concert_id).collection("setlists").document(track_doc_id).set(setlist_data)
        print(f"Uploaded setlist for concert: {concert_name}, track: {title}")

upload_excel_to_Firestore("C:/Users/qsc13/Documents/day6_data/day6_concert_data.xlsx")
